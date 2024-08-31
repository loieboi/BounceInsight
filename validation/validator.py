import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt
from scipy.stats import pearsonr
import numpy as np
import os


class Validator:

    def __init__(self, tolerance=0.05):
        self.tolerance = tolerance

    def validate(self, tolerance=0.05):
        df_fp, df_gym = self.load_validator_files()
        self.validate_data(df_fp, df_gym, tolerance=tolerance)
        self.plot_bland_altman(df_fp, df_gym)
        self.calculate_correlations(df_fp, df_gym)

    def load_validator_files(self):
        df_fp = pd.read_csv('validation/validation_forceplate.csv', dtype={'participant_id': str})
        df_gym = pd.read_csv('validation/validation_gymaware.csv', dtype={'participant_id': str})
        return df_fp, df_gym

    def validate_data(self, df_fp, df_gym, tolerance):
        merged_df = pd.merge(df_fp, df_gym, on=['file_name', 'participant_id'], suffixes=('_fp', '_gym'))

        columns_to_compare = {
            't_ecc': 't_ecc',
            't_con': 't_con',
            't_total': 't_total',
            'F_turning': 'F_turning',
            'pFc': 'pFc',
            'mFc': 'mFc'
        }

        validation_results = []

        for _, row in merged_df.iterrows():
            file_name = row['file_name']
            participant_id = row['participant_id']
            result = {
                'file_name': file_name,
                'participant_id': participant_id,
                't_ecc_diff': None,
                't_ecc_comparison': None,
                't_con_diff': None,
                't_con_comparison': None,
                't_total_diff': None,
                't_total_comparison': None,
                'F_turning_diff': None,
                'F_turning_comparison': None,
                'pFc_diff': None,
                'pFc_comparison': None,
                'mFc_diff': None,
                'mFc_comparison': None,
            }

            for col_fp, col_gym in columns_to_compare.items():
                try:
                    if col_fp in row and col_gym in row:
                        fp_value = row[col_fp]
                        gym_value = row[col_gym]
                        diff = abs(fp_value - gym_value)
                        if diff > tolerance * max(fp_value, gym_value):
                            result[f'{col_fp}_diff'] = round(diff, 2)
                            result[f'{col_fp}_comparison'] = 'fp_higher' if fp_value > gym_value else 'gym_higher'
                    elif f'{col_fp}_fp' in row and f'{col_gym}_gym' in row:
                        fp_value = row[f'{col_fp}_fp']
                        gym_value = row[f'{col_gym}_gym']
                        diff = abs(fp_value - gym_value)
                        if diff > tolerance * max(fp_value, gym_value):
                            result[f'{col_fp}_diff'] = round(diff, 2)
                            result[f'{col_fp}_comparison'] = 'fp_higher' if fp_value > gym_value else 'gym_higher'
                except KeyError as e:
                    print(f"KeyError: {e}")

            validation_results.append(result)

        validation_df = pd.DataFrame(validation_results)

        validation_df = validation_df.dropna(subset=['t_ecc_diff', 't_con_diff', 't_total_diff', 'F_turning_diff'],
                                             how='all')

        validation_df.to_csv('validation/validation_results.csv', index=False)
        print('Results saved to validation/validation_results.csv')

        self.create_excel(validation_df)
        print('Validation complete')

    def create_excel(self, validation_df):
        excel_path = 'validation/validation_results.xlsx'
        validation_df.to_excel(excel_path, index=False)
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.column_letter in ['C', 'E', 'G']:  # t_ecc_diff, t_con_diff, t_total_diff
                    try:
                        if cell.value is not None:
                            if (cell.column_letter == 'C' or cell.column_letter == 'E') and cell.value < 0.75:
                                cell.fill = green_fill
                            elif (cell.column_letter == 'C' or cell.column_letter == 'E') and cell.value > 0.75:
                                cell.fill = red_fill
                            elif cell.column_letter == 'G' and cell.value < 1:
                                cell.fill = green_fill
                            elif cell.column_letter == 'G' and cell.value > 1:
                                cell.fill = red_fill
                    except TypeError:
                        pass
                elif cell.column_letter in ['I', 'K', 'M']:  # F_turning_diff, pFc_diff, mFc_diff
                    try:
                        if cell.value is not None:
                            if cell.value < 750:
                                cell.fill = green_fill
                            elif cell.value > 750:
                                cell.fill = red_fill
                    except TypeError:
                        pass

        wb.save(excel_path)
        print(f'Visualized validation results saved to {excel_path}')

    def plot_bland_altman(self, df_fp, df_gym):
        # Merge the dataframes on 'file_name' and 'participant_id'
        merged_df = pd.merge(df_fp, df_gym, on=['file_name', 'participant_id'], suffixes=('_fp', '_gym'))

        # Comparisons for time measurements
        ecc_comparisons = [
            ('t_ecc_fp', 't_ecc_gym', 't_ecc')
        ]

        con_comparisons = [
            ('t_con_fp', 't_con_gym', 't_con')
        ]

        time_comparisons = [
            ('t_total_fp', 't_total_gym', 't_total')
        ]

        # Comparisons for force measurements
        force_comparisons = [
            ('F_turning_fp', 'F_turning_gym', 'F_turning')

        ]

        peak_force_con_comparisons = [
            ('pFc_fp', 'pFc_gym', 'pFc')
        ]
        mean_force_con_comparisons = [
            ('mFc_fp', 'mFc_gym', 'mFc')
        ]

        # Create Bland-Altman plot for time measurements
        self._bland_altman_plot(merged_df, ecc_comparisons, "Bland-Altman Plot for Eccentric Time Measurements", color='green')
        self._bland_altman_plot(merged_df, con_comparisons, "Bland-Altman Plot for Concentric Time Measurements", color='blue')
        self._bland_altman_plot(merged_df, time_comparisons, "Bland-Altman Plot for Total Time Measurements", color='red')

        # Create Bland-Altman plot for force measurements
        self._bland_altman_plot(merged_df, force_comparisons, "Bland-Altman Plot for Turning Force Measurements", color='green')

        # Create Bland-Altman plot for Concentric force measurements
        self._bland_altman_plot(merged_df, peak_force_con_comparisons,
                                "Bland-Altman Plot for Peak Concentric Force Measurements", color='blue')
        self._bland_altman_plot(merged_df, mean_force_con_comparisons,
                                "Bland-Altman Plot for Mean Concentric Force Measurements", color='blue')

    def _bland_altman_plot(self, df, comparisons, title, color='blue'):
        plt.figure(figsize=(12, 8))

        # Define different colors for each comparison
        colors = color

        overall_diffs = []

        for i, (col1, col2, label) in enumerate(comparisons):
            mean = (df[col1] + df[col2]) / 2
            diff = df[col1] - df[col2]
            overall_diffs.extend(diff.dropna())  # Ensure no NaN values are included
            plt.scatter(mean, diff, alpha=0.5, label=label, color=colors[i % len(colors)])

        # Calculate overall mean difference and standard deviation
        if overall_diffs:
            overall_diffs = np.array(overall_diffs)
            mean_diff = np.mean(overall_diffs)
            std_diff = np.std(overall_diffs)

            # Plot overall mean difference and limits of agreement
            plt.axhline(mean_diff, color='gray', linestyle='--', label='Mean Difference')
            plt.axhline(mean_diff + 1.96 * std_diff, color='red', linestyle='--', label='Limits of Agreement')
            plt.axhline(mean_diff - 1.96 * std_diff, color='red', linestyle='--')

            # Evaluate the points within and outside the limits of agreement
            lower_limit = mean_diff - 1.96 * std_diff
            upper_limit = mean_diff + 1.96 * std_diff

            outside_limits = np.sum((overall_diffs < lower_limit) | (overall_diffs > upper_limit))
            within_limits = np.sum((overall_diffs >= lower_limit) & (overall_diffs <= upper_limit))

            total_points = len(overall_diffs)
            print(f"Total points: {total_points}")
            print(f"Points within limits: {within_limits} ({within_limits / total_points * 100:.2f}%)")
            print(f"Points outside limits: {outside_limits} ({outside_limits / total_points * 100:.2f}%)")

            # Determine if the data is good or not
            if within_limits / total_points > 0.95:
                print("The data is good (more than 95% points within limits of agreement).")
            else:
                print("The data is not good (less than 95% points within limits of agreement).")

            print(f'Mean {mean_diff}, STD {std_diff}')
            print(f'Lower limit: {lower_limit}, Upper limit: {upper_limit}, Span: {upper_limit - lower_limit}')

        plt.title(title)
        plt.xlabel('Mean of Two Measurements')
        plt.ylabel('Difference between Two Measurements')
        plt.legend()
        plt.grid(True)
        plt.show()

    def calculate_correlations(self, df_fp, df_gym):
        merged_df = pd.merge(df_fp, df_gym, on=['file_name', 'participant_id'], suffixes=('_fp', '_gym'))
        comparisons = [
            ('t_ecc_fp', 't_ecc_gym', 't_ecc'),
            ('t_con_fp', 't_con_gym', 't_con'),
            ('t_total_fp', 't_total_gym', 't_total'),
            ('F_turning_fp', 'F_turning_gym', 'F_turning'),
            ('pFc_fp', 'pFc_gym', 'pFc'),
            ('mFc_fp', 'mFc_gym', 'mFc')
        ]
        for col1, col2, label in comparisons:
            valid_data = merged_df[[col1, col2]].dropna()
            correlation, p_value = pearsonr(valid_data[col1], valid_data[col2])
            if p_value < 0.001:
                p_str = "< 0.001"
            elif p_value < 0.005:
                p_str = "< 0.005"
            elif p_value < 0.05:
                p_str = "< 0.05"
            else:
                p_str = f"= {p_value:.3f}"
            print(f"Correlation for {label}: {correlation:.2f}, p-value {p_str}")




